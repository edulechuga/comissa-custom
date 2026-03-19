import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import re

from execution.calculo_impostos import CalculadoraImpostos

try:
    import fitz
except ImportError:
    fitz = None
    print("PyMuPDF (fitz) não instalado. Leitura de NF não funcionará.")


class GeradorComissao:
    def __init__(self, db_path: str):
        self.calculadora = CalculadoraImpostos(db_path)

    def ler_faturas_e_numero_nf(self, caminho_nf: str) -> tuple[list, str]:
        """ Extrai as datas exclusivas numéricas da Fatura e o Numero da NF """
        if not caminho_nf or not os.path.exists(caminho_nf) or not fitz:
            return [], "XXXXX"
        try:
            doc = fitz.open(caminho_nf)
            text = ""
            for p in doc:
                text += p.get_text()
            
            text_upper = text.upper()
            idx_fatura = text_upper.find('FATURA')
            idx_imposto = text_upper.find('CÁLCULO DO IMPOSTO')
            if idx_imposto == -1: 
                idx_imposto = text_upper.find('CALCULO DO IMPOSTO')
                
            if idx_fatura != -1:
                # Extrai apenas text que vem após a palavra FATURA e antes do CALCULO DO IMPOSTO
                bloco_fatura = text[idx_fatura:idx_imposto] if idx_imposto != -1 else text[idx_fatura:idx_fatura+600]
                dates = re.findall(r'\d{2}/\d{2}/\d{4}', bloco_fatura)
            else:
                # Fallback genérico se o PDF tiver formato diferente
                dates = re.findall(r'\d{2}/\d{2}/\d{4}', text[text_upper.find('FATURA'):] if 'FATURA' in text_upper else text)
            
            # Filtra pro formato de fatura (removendo as repetições)
            unique_dates = []
            for d in set(dates):
                # Usar um sort preservando a extração original das datas, ou deixá-la em ordem cronológica com um parsing simples
                pass
                
            # As datas na NF geralmente estão em ordem
            unique_dates = sorted(list(set(dates)), key=lambda x: (x.split('/')[2], x.split('/')[1], x.split('/')[0]))
            
            # Buscar número da NF "N. 000033515"
            nf_match = re.search(r'N\.\s*0*(\d+)', text)
            nf_num = str(nf_match.group(1))[-5:] if nf_match else "XXXXX"
            
            return unique_dates, nf_num
        except Exception as e:
            print(f"Erro lendo NF PDF: {e}")
            return [], "XXXXX"

    def extrair_custos(self, caminho_pedido: str) -> dict:
        df = pd.read_excel(caminho_pedido, nrows=100)
        custos = {}
        flag = False
        col_pn, col_custo = 1, 3

        for i, row in df.iterrows():
            vals = [str(v).strip().upper() for v in row.values if pd.notna(v)]
            if "CUSTO REVENDA PARA EFEITOS DE COMISSIONAMENTO" in vals:
                flag = True
                continue
                
            if flag and "CÓDIGO DO PRODUTO" in vals and "CUSTO UNITÁRIO" in vals:
                for idx, v in enumerate(row.values):
                    v_str = str(v).strip().upper() if pd.notna(v) else ""
                    if v_str == "CÓDIGO DO PRODUTO": col_pn = idx
                    if v_str == "CUSTO UNITÁRIO": col_custo = idx
                continue
                
            if flag:
                if len(row.values) > max(col_pn, col_custo):
                    pn = str(row.values[col_pn]).strip()
                    custo = row.values[col_custo]
                    
                    if pn and pn != "nan" and pn != "0" and pd.notna(custo):
                        try:
                            custos[pn] = float(custo)
                        except:
                            pass
        return custos

    def gerar_planilha_blocos(self, caminho_pedido: str, caminho_nf_pdf: str, caminho_saida_base_dir: str) -> str:
        print("Mapeando impostos pela Camada 1...")
        metadados = self.calculadora.ler_pedido_venda(caminho_pedido)
        resultados_fiscais = self.calculadora.processar_pedido_completo(caminho_pedido)
        
        print("Mapeando tabela de custos...")
        custos_pn = self.extrair_custos(caminho_pedido)
        
        print("Mapeando Faturas da NF...")
        faturas_datas, num_nf = self.ler_faturas_e_numero_nf(caminho_nf_pdf) if caminho_nf_pdf else ([], "XXXXX")
        
        razao_social = metadados.get("RAZAO_SOCIAL", "Desconhecida")
        # HIGIENIZAR a razão social para não quebrar a criação de arquivo no SO Windows/Mac:
        razao_social = re.sub(r'[^A-Za-z0-9 _-]', '', razao_social).strip()
        razao_social = re.sub(r'\s+', ' ', razao_social)
        
        # Formata NOME DA SAIDA de acordo com PDF e Pedido
        nome_arquivo_final = f"COMISSAO - NF {num_nf} {razao_social}.xlsx"
        caminho_saida = os.path.join(caminho_saida_base_dir, nome_arquivo_final)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        estado = metadados.get("ESTADO_DESTINO", "XX")
        ws.title = f"{estado} - COMISSIONAMENTO"
        
        taxa_pis = 0.0165
        taxa_cofins = 0.076
        col_offset = 1
        
        nat_op = metadados.get("NATUREZA_OPERACAO", "CONSUMO")
        ie_txt = "ISENTO" if metadados.get("TEM_IE", "NÃO") == "NÃO" else "CONTRIBUINTE"
        
        # Paleta de cores estilo original do Excel
        fill_header = PatternFill(fill_type="solid", fgColor="DDEBF7") # Azul claro
        fill_dark = PatternFill(fill_type="solid", fgColor="8EA9DB") # Azul escurinho
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for item in resultados_fiscais:
            pn = item['PN']
            desc = item['Descricao']
            venda = float(item['Preco_Unitario'])
            qtd_vendida = int(item['Quantidade']) if str(item['Quantidade']).isdigit() else float(item['Quantidade']) if item['Quantidade'] else 1
            custo = custos_pn.get(pn, 0.0)
            
            aliq_icms_base = float(item['Aliq_Interestadual_Calculada'])
            aliq_ipi = float(item.get('Aliq_IPI', 0.0))
            
            imposto_calculado = item['Imposto_Final']
            tipo_imposto = item.get('Tipo_Imposto', '')
            
            try:
                val_imposto_calculado = float(imposto_calculado)
            except (ValueError, TypeError):
                val_imposto_calculado = 0.0
            
            
            # Lógica de Roteamento de onde colocar o valor da Taxa (Diferença de Alíquota ou Difal)
            # O sistema vai puxar da engine no formato Tipo = "DIFAL", "ICMS ST ", "ERRO"
            taxa_difal_contribuinte = val_imposto_calculado if tipo_imposto == "DIFAL" else 0.0
            taxa_dif_aliquota_st = val_imposto_calculado if tipo_imposto == "ICMS ST " else 0.0
            
            # --- DESENHANDO O BLOCO NO EXCEL ---
            col_a = get_column_letter(col_offset)
            col_b = get_column_letter(col_offset+1)
            col_c = get_column_letter(col_offset+2)
            col_d = get_column_letter(col_offset+3)
            col_e = get_column_letter(col_offset+4)
            
            # Titulos do bloco do Item (Mesclado como no template)
            ws.merge_cells(f"{col_a}2:{col_e}2")
            cell_titulo = ws.cell(row=2, column=col_offset, value=f"{nat_op} {ie_txt} - {estado}")
            cell_titulo.font = Font(bold=True, size=14)
            cell_titulo.alignment = Alignment(horizontal="center")
            cell_titulo.fill = fill_header
            
            ws.merge_cells(f"{col_a}5:{col_e}6") # Produto Box
            cell_prod = ws.cell(row=5, column=col_offset, value=f"{pn} - {desc}")
            cell_prod.font = Font(bold=True, size=11, color="FFFFFF")
            cell_prod.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_prod.fill = fill_dark
            
            # Linha CUSTO PROJETOS / VENDA
            ws.cell(row=8, column=col_offset, value="CUSTO PROJETOS").font = Font(bold=True)
            ws.cell(row=8, column=col_offset+1, value="VENDA").font = Font(bold=True)
            ws.cell(row=8, column=col_offset+2, value="Diferença").font = Font(bold=True)
            ws.cell(row=8, column=col_offset+3, value="Frete ").font = Font(bold=True)
            ws.cell(row=8, column=col_offset+4, value="Prazo").font = Font(bold=True)
            
            # Valores 
            for c_idx in range(col_offset, col_offset+3):
                ws.cell(row=9, column=c_idx).number_format = '#,##0.00'
                
            ws.cell(row=9, column=col_offset, value=custo)
            ws.cell(row=9, column=col_offset+1, value=venda)
            # Diferenca = Custo - Venda (B9-A9) -> traduzido pro dinamico
            ws.cell(row=9, column=col_offset+2, value=f"={col_b}9-{col_a}9")
            
            # Diferença Aliquota (Para casos ICMS ST)
            ws.cell(row=12, column=col_offset, value="Diferença Aliquota (ICMS ST)")
            ws.cell(row=12, column=col_offset+1, value=taxa_dif_aliquota_st).number_format = '0.00%'
            ws.cell(row=12, column=col_offset+2, value=f"={col_c}9-({col_c}9/(1+{col_b}12))").number_format = '#,##0.00'
            
            # Difal (Para casos Difal)
            ws.cell(row=13, column=col_offset, value="Difal")
            ws.cell(row=13, column=col_offset+1, value=taxa_difal_contribuinte).number_format = '0.00%'
            ws.cell(row=13, column=col_offset+2, value=f"={col_c}9*{col_b}13").number_format = '#,##0.00'
            
            # Comissoes 
            ws.cell(row=15, column=col_offset, value="Comissao Sem Dif Aliquota")
            ws.cell(row=15, column=col_offset+2, value=f"={col_c}9-{col_c}12").number_format = '#,##0.00'
            
            ws.cell(row=16, column=col_offset, value="Comissao Sem Difal")
            ws.cell(row=16, column=col_offset+2, value=f"=0").number_format = '#,##0.00'
            
            # Comissao bruta:
            ws.cell(row=18, column=col_offset, value="Comissão bruta").font = Font(bold=True)
            ws.cell(row=18, column=col_offset+2, value=f"=({col_c}16-{col_c}15*-1)-{col_c}13").number_format = '#,##0.00'
            
            # Cascatas Impostos
            ws.cell(row=19, column=col_offset, value="IPI")
            ws.cell(row=19, column=col_offset+1, value=aliq_ipi).number_format = '0.00%'
            ws.cell(row=19, column=col_offset+2, value=f"={col_c}18-{col_c}18/(1+{col_b}19)").number_format = '#,##0.00'
            
            ws.cell(row=20, column=col_offset, value="PIS")
            ws.cell(row=20, column=col_offset+1, value=taxa_pis).number_format = '0.00%'
            ws.cell(row=20, column=col_offset+2, value=f"=({col_c}18-{col_c}19)*{col_b}20").number_format = '#,##0.00'
            
            ws.cell(row=21, column=col_offset, value="COFINS")
            ws.cell(row=21, column=col_offset+1, value=taxa_cofins).number_format = '0.00%'
            ws.cell(row=21, column=col_offset+2, value=f"=({col_c}18-{col_c}19)*{col_b}21").number_format = '#,##0.00'

            ws.cell(row=22, column=col_offset, value="ICMS")
            ws.cell(row=22, column=col_offset+1, value=aliq_icms_base).number_format = '0.00%'
            ws.cell(row=22, column=col_offset+2, value=f"=({col_c}18-{col_c}19)*{col_b}22").number_format = '#,##0.00'

            # Final - Comissão Líquida + Adicionado Venda Total em E23/D23 Demandada pelo Usuário
            cel_cl = ws.cell(row=23, column=col_offset, value="Comissão liquida Un.")
            cel_cl.font = Font(bold=True, size=12)
            cel_cl.fill = PatternFill(fill_type="solid", fgColor="FFFF00") # Amarelinho
            
            # Comissão Líquida Uni (Calculo)
            ws.cell(row=23, column=col_offset+2, value=f"={col_c}18/(1+{col_b}19)*(1-{col_b}20-{col_b}21-{col_b}22)").font = Font(bold=True)
            ws.cell(row=23, column=col_offset+2).number_format = '#,##0.00'
            
            # Quantidade (Coluna D)
            ws.cell(row=23, column=col_offset+3, value=qtd_vendida).font = Font(bold=True, color="FF0000")
            ws.cell(row=23, column=col_offset+3).alignment = Alignment(horizontal="center")
            
            # Total Comissão Liquida (Comissão * Qtd) (Coluna E)
            cel_clt = ws.cell(row=23, column=col_offset+4, value=f"={col_c}23*{col_d}23")
            cel_clt.font = Font(bold=True, size=12)
            cel_clt.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
            cel_clt.number_format = '#,##0.00'
            
            # --- INCORPORANDO AS FATURAS NA PARTE DE BAIXO (EX: D24, E24...) ---
            if faturas_datas:
                qtd_faturas = len(faturas_datas)
                start_faturas_row = 24
                # Estilo pro cabecalho
                ws.cell(row=start_faturas_row-1, column=col_offset, value="Previsao Parcelas (NF):").font = Font(italic=True, color="0000FF")
                
                for idx_fat, data_fat in enumerate(faturas_datas):
                    linha_fat = start_faturas_row + idx_fat
                    
                    # Célula de Data (Na Coluna D onde seria D24)
                    ws.cell(row=linha_fat, column=col_offset+3, value=data_fat).alignment = Alignment(horizontal="center")
                    
                    # Célula de Divisao Parcela (Na Coluna E, usando FÓRMULA total_liquido / len)
                    # ex: =E23/3
                    ws.cell(row=linha_fat, column=col_offset+4, value=f"={col_e}23/{qtd_faturas}").number_format = '#,##0.00'
                    ws.cell(row=linha_fat, column=col_offset+4).fill = PatternFill(fill_type="solid", fgColor="E2EFDA") # Verde claro parcelas
                    
            # Larguras esteticas pro bloco atual
            ws.column_dimensions[col_a].width = 30
            ws.column_dimensions[col_b].width = 12
            ws.column_dimensions[col_c].width = 16
            ws.column_dimensions[col_d].width = 15
            ws.column_dimensions[col_e].width = 16
            
            for border_row in range(8, 24 + len(faturas_datas)):
                for c_bd in range(col_offset, col_offset+5):
                    ws.cell(row=border_row, column=c_bd).border = thin_border
            
            # Pula 6 colunas para o proximo produto do array
            col_offset += 6

        wb.save(caminho_saida)
        print(f"\\n(OK) Planilha gerada com sucesso e salva em: {caminho_saida}")
        return caminho_saida

if __name__ == "__main__":
    caminho_base = os.path.dirname(os.path.dirname(__file__))
    impostos_file = os.path.join(caminho_base, "IMPOSTOS.xlsx")
    ncm_file = os.path.join(caminho_base, "NCM.xlsx")
    pedido_file = os.path.join(caminho_base, "RQCM-001 - Pedido de venda - O.LIVE - 02022026.xlsx")
    saida = os.path.join(caminho_base, ".tmp", "Gerado_Calculo_Comissao.xlsx")
    
    # PDF Opcional
    nf_file = os.path.join(caminho_base, "33515.pdf")
    if not os.path.exists(nf_file):
        nf_file = None
        
        # Passa o tmp dir no teste isolado
        db_path = os.path.join(caminho_base, ".tmp", "impostos.db")
        from execution.database_manager import initialize_database
        initialize_database(impostos_file, ncm_file, db_path)
        
        gerador = GeradorComissao(db_path)
        tmp_dir = os.path.join(caminho_base, ".tmp")
        gerador.gerar_planilha_blocos(pedido_file, nf_file, tmp_dir)
    else:
        print("Arquivos locais incompletos.")
